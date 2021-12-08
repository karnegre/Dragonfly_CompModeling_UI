import streamlit as st
import os
import os.path
import pandas as pd
from openpyxl import load_workbook
from csv import writer
# ------------------------------------------------------
def app():

    cwd =os.getcwd()
    fpath_oc= os.path.join(cwd,'userupload','oc')
    fpath_ir= os.path.join(cwd,'userupload','ir')
    fpath_user= os.path.join(cwd,'userupload','userinfo')


    # def userinfotoexcel(df, cwd):
    #     path_exel = os.path.join(cwd,'userupload','uexcel.xlsx')
    #     book = load_workbook(path_exel)
    #     writer = pd.ExcelWriter("uexcel.xlsx", engine='openpyxl')
    #     writer.book = book
    #     writer.sheets = {ws.title: ws for ws in book.worksheets}
    #     df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)
    #     writer.save()
    #     return st.success("Information added")

    def save_uploaded_OCfile(upload):
        with open(os.path.join(fpath_oc,upload.name),"wb") as f:
            check = st. radio("Is your data file in the correct format?",('No','Yes'))
            if check == "Yes":
                f.write(upload.getbuffer())
                return st.success("Saved file : {} in fpath_oc".format(upload.name))

    def save_uploaded_IRfile(upload):
        with open(os.path.join(fpath_ir,upload.name),"wb") as f:
            check = st. radio("Is your data file in the correct format?",('No','Yes'))
            if check == "Yes":
                f.write(upload.getbuffer())
                return st.success("Saved file : {} in fpath_ir".format(upload.name))

    st.header('Uploading Data to the App')
    st.subheader('User Information')
    df_user=pd.DataFrame(columns=['Name','Email','DOI',"Upload date", "Data Type","File Uploaded"])
    name = st.text_input("Enter your name:")
    email = st.text_input("Enter your email:")
    DOI = st.text_input("Enter the DOI:")
    date= st.date_input("Select todays date")

    dataset= ["OC", "IR"]
    choice = st.radio("Data Upload Type", dataset)

    if choice == "OC":
        st.subheader("Optical Constant Data Upload")
        datafile = st.file_uploader("Upload ", type= ['txt'])
        if datafile is not None:
            filedetails = {"FileName": datafile.name, "FileType": datafile.type}
            df = pd.read_csv(datafile,sep = "\t")
            df_user.loc[len(df_user.index)] = [name,email, DOI, date, choice, datafile.name ]
            st.write(df_user)
            st.dataframe(df)

            path_csv = os.path.join(cwd,'userupload','userinfo.csv')
            with open (path_csv, 'a', newline='') as f_obj:
                df_user.to_csv(f_obj, sep = "\t", header=False, index = False)
            #
            # with pd.ExcelWriter("UserInfo.xlsx", mode="a", engine="openpyxl") as writer:
            #     df_user.to_excel(writer, startrow=writer.sheets['info'].max_row, index = False,header= False)

            save_uploaded_OCfile(datafile)

    elif choice == "IR":
        st.subheader("Reflectance Spectra Data Upload")
        datafile = st.file_uploader("Upload ", type= ['txt'])
        if datafile is not None:
            filedetails = {"FileName": datafile.name, "FileType": datafile.type}
            df = pd.read_csv(datafile,sep = "\t")
            df_user.loc[len(df_user.index)] = [name,email, DOI, date, choice, datafile.name ]
            st.write(df_user)
            st.dataframe(df)

            path_csv = os.path.join(cwd,'userupload','userinfo.csv')
            with open (path_csv, 'a', newline='') as f_obj:
                df_user.to_csv(f_obj, sep = "\t", header=False, index = False)
                
            save_uploaded_IRfile(datafile)
